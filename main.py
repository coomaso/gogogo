import requests
import base64
import json
from Crypto.Cipher import AES
import time
from urllib.parse import quote
import random
import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, Color
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import WorkbookProperties

# 配置常量
HEADERS = {
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,vi;q=0.7",
    "Connection": "keep-alive",
    "Content-Type": "application/json; charset=utf-8",
    "Host": "106.15.60.27:22222",
    "Referer": "http://106.15.60.27:22222/xxgs/",
    "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.95 Safari/537.36"
}

RETRY_COUNT = 3               # 请求重试次数
PAGE_RETRY_MAX = 2           # 单页最大重试次数
TIMEOUT = 15                  # 请求超时时间（秒）
PAGE_SIZE = 10

# AES配置
AES_KEY = b"6875616E6779696E6875616E6779696E"
AES_IV = b"sskjKingFree5138"

def safe_request(session: requests.Session, url: str) -> requests.Response:
    """带自动重试的安全请求"""
    for attempt in range(RETRY_COUNT):
        try:
            if attempt > 0:
                time.sleep(random.uniform(0.5, 2.5))
            print(f"正在请求: {url}")  # 添加请求URL日志
            response = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            response.raise_for_status()
            return response
        except requests.exceptions.Timeout:
            print(f"↺ 请求超时，正在重试 ({attempt+1}/{RETRY_COUNT})...")
        except requests.exceptions.RequestException as e:
            print(f"请求异常: {str(e)}")  # 打印具体异常信息
            if attempt < RETRY_COUNT - 1:
                print(f"正在进行第 {attempt+2} 次尝试...")
    raise RuntimeError(f"超过最大重试次数 ({RETRY_COUNT})")

def aes_decrypt_base64(encrypted_base64: str) -> str:
    """增强版AES解密函数"""
    if not encrypted_base64:
        raise ValueError("加密数据为空，无法解密")

    try:
        encrypted_bytes = base64.b64decode(encrypted_base64)
        cipher = AES.new(AES_KEY, AES.MODE_CBC, AES_IV)
        decrypted_bytes = cipher.decrypt(encrypted_bytes)
        return decrypted_bytes.rstrip(b'\x00').decode("utf-8")
    except Exception as e:
        print(f"解密失败，原始数据: {encrypted_base64[:50]}...")  # 打印部分原始数据
        raise RuntimeError(f"解密失败: {str(e)}")

def get_new_code(session: requests.Session) -> tuple:
    """获取新验证码和时间戳"""
    timestamp = str(int(time.time() * 1000))
    code_url = f"http://106.15.60.27:22222/ycdc/bakCmisYcOrgan/getCreateCode?codeValue={timestamp}"

    try:
        response = safe_request(session, code_url).json()
        print(f"验证码接口响应: {json.dumps(response, ensure_ascii=False)[:100]}...")  # 打印部分响应
        if response.get("code") != 0:
            raise RuntimeError(f"验证码接口异常: {response}")
        return aes_decrypt_base64(response["data"]), timestamp
    except Exception as e:
        print(f"获取验证码失败，URL: {code_url}")  # 打印失败的URL
        raise RuntimeError(f"获取新验证码失败: {str(e)}")

def parse_response_data(encrypted_data: str) -> dict:
    """健壮的数据解析方法"""
    if not encrypted_data:
        print("警告: 收到空的加密数据")  # 添加警告日志
        return {"error": "empty data"}

    try:
        decrypted_str = aes_decrypt_base64(encrypted_data)
        print(f"解密后的数据样本: {decrypted_str[:100]}...")  # 打印解密后的数据样本
        return json.loads(decrypted_str)
    except json.JSONDecodeError as e:
        print(f"JSON解析错误，数据样本: {encrypted_data[:50]}...")  # 打印错误数据样本
        return {"error": f"invalid json format: {str(e)}"}
    except Exception as e:
        return {"error": str(e)}

def process_page(session: requests.Session, page: int, code: str, timestamp: str) -> tuple:
    """处理单个页面并返回数据，包含重试机制"""
    max_retries = 3
    current_code = code
    current_timestamp = timestamp

    for attempt in range(max_retries + 1):
        page_url = (
            "http://106.15.60.27:22222/ycdc/bakCmisYcOrgan/getCurrentIntegrityPage"
            f"?pageSize={PAGE_SIZE}&cioName=%E5%85%AC%E5%8F%B8&page={page}"
            f"&code={quote(current_code)}&codeValue={current_timestamp}"
        )

        try:
            # 发送带当前参数的请求
            response = safe_request(session, page_url)
            page_response = response.json()
            status = page_response.get('code', '未知')
            print(f"第 {page} 页 请求#{attempt+1} 响应状态: {status}")

            # 空数据检查
            if "data" not in page_response or not page_response["data"]:
                print(f"空数据响应，准备重试（剩余重试次数: {max_retries - attempt}）")
                if attempt < max_retries:# 重试三次
                    response = safe_request(session, page_url)
                    page_response = response.json()
                    continue
                raise RuntimeError("连续空响应，终止重试")

            # 数据解析
            page_data = parse_response_data(page_response["data"])

            records = page_data.get("data", [])
            print(f"第 {page} 页解析出 {len(records)} 条记录")  # 明确记录数量
            
            # 检查解析出的数据是否有效
            if not records:
                print(f"警告: 第 {page} 页解析出空记录列表")
                
            return records, page_data.get("total", 0)
        except Exception as e:
            print(f"第 {page} 页处理失败: {str(e)}")
            raise

    raise RuntimeError("超过最大重试次数")

def fetch_company_detail(session: requests.Session, cec_id: str, company_name: str, max_retries=3) -> dict:
    """获取企业信誉分明细（增强版，带重试）"""
    print(f"\n获取企业信誉分明细: {company_name} (cecId={cec_id})")
    detail_url = f"http://106.15.60.27:22222/ycdc/bakCmisYcOrgan/getCurrentIntegrityDetails?cecId={cec_id}"
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            response = safe_request(session, detail_url)
            response_data = response.json()

            if response_data.get("code") != 0:
                print(f"信誉分明细接口异常: {response_data}")
                last_error = f"接口异常: {response_data}"
                continue

            encrypted_data = response_data.get("data", "")
            if not encrypted_data:
                print("信誉分明细接口返回空数据")
                last_error = "接口返回空数据"
                continue
            parse_response_data(encrypted_data)
            decrypted_str = aes_decrypt_base64(encrypted_data)
            detail_data = json.loads(decrypted_str)

            company_detail = {
                "cioName": detail_data.get("data", {}).get("cioName", company_name),
                "jfsj": detail_data.get("data", {}).get("jfsj", ""),
                "eqtName": detail_data.get("data", {}).get("eqtName", ""),
                "blxwArray": detail_data.get("data", {}).get("blxwArray", []),
                "lhxwArray": detail_data.get("data", {}).get("lhxwArray", []),
                "cecId": detail_data.get("data", {}).get("cecId", cec_id),
                "cechId": detail_data.get("data", {}).get("cechId", "")
            }
            print(f"成功获取企业信誉分明细: {company_detail.get('cioName')}")
            print(company_detail)
            return company_detail

        except Exception as e:
            print(f"第{attempt}次获取企业信誉分明细失败: {str(e)}")
            last_error = str(e)
            time.sleep(random.uniform(1, 5))  # 延迟后重试

    print(f"获取企业信誉分明细失败: {last_error}")
    return {}

def append_top_json(sorted_data, category_name, github_mode=False):
    """追加数据到当天的JSON文件"""
    # 获取当前日期（北京时间）
    utc8_offset = timezone(timedelta(hours=8))
    now = datetime.now(utc8_offset)
    date_str = now.strftime("%Y%m%d")
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    
    # 确定输出目录
    output_dir = os.getcwd()
    if github_mode:
        output_dir = os.path.join(output_dir, "excel_output")
        os.makedirs(output_dir, exist_ok=True)
    
    # 构建固定日期的文件名
    json_filename = f"{category_name}_top10.json"
    json_path = os.path.join(output_dir, json_filename)
    
    # 准备本次数据
    data_list = []
    for idx, item in enumerate(sorted_data[:10], 1):
        # 添加企业基本信息
        company_data = {
            "排名": idx,
            "企业名称": item.get("cioName", ""),
            "诚信分值": item.get("score", 0),
            "组织ID": item.get("cecId", ""),
        }
        
        # 添加信誉分明细
        if "detail" in item:
            company_data["信誉分明细"] = item["detail"]
        
        data_list.append(company_data)
    
    # 构建本次更新数据
    update_data = {
        "TIMEamp": timestamp,
        "DATAlist": data_list
    }
    
    # 读取或初始化JSON文件
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
                
            # 确保数据结构正确
            if not isinstance(existing_data, list):
                # 如果是旧格式，转换为新格式
                existing_data = [existing_data]
                
            existing_data.append(update_data)
        except:
            # 文件损坏时重新开始
            existing_data = [update_data]
    else:
        existing_data = [update_data]
    
    # 写入更新后的数据
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)
        print(f"已追加数据到JSON文件: {os.path.abspath(json_path)}")
        return json_path
    except Exception as e:
        print(f"JSON文件追加失败: {str(e)}")
        return None

def export_to_excel(data, session, github_mode=False):
    """
    专业级Excel导出函数（多工作表分类排序）
    :param data: 待导出数据列表，每个元素为字典格式
    :param session: 请求会话
    :param github_mode: 是否启用GitHub Actions模式
    :return: 生成的Excel文件绝对路径
    """
    # ==================== 列配置 ====================
    COLUMNS = [
        {'id': 'cioName',    'name': '企业名称',   'width': 35,  'merge': True,  'align': 'left'},
        {'id': 'eqtName',    'name': '资质类别',   'width': 20,  'merge': True,  'align': 'center'},
        {'id': 'csf',        'name': '初始分',     'width': 12,  'merge': True,  'align': 'center', 'format': '0'},
        {'id': 'zzmx',       'name': '资质明细',   'width': 50,  'merge': False, 'align': 'left'},
        {'id': 'cxdj',       'name': '诚信等级',   'width': 12,  'merge': False, 'align': 'center'},
        {'id': 'score',      'name': '诚信分值',   'width': 12,  'merge': False, 'align': 'center', 'format': '0.0'},
        {'id': 'jcf',        'name': '基础分',     'width': 12,  'merge': False, 'align': 'center', 'format': '0'},
        {'id': 'zxjf',       'name': '专项加分',   'width': 12,  'merge': False, 'align': 'center', 'format': '0.0'},
        {'id': 'kf',         'name': '扣分',       'width': 12,  'merge': False, 'align': 'center', 'format': '0.0'},
        {'id': 'eqlId',      'name': '资质ID',     'width': 25,  'merge': False, 'align': 'center'},
        {'id': 'orgId',      'name': '组织ID',     'width': 30,  'merge': True,  'align': 'center'},
        {'id': 'cecId',      'name': '信用档案ID', 'width': 30,  'merge': True,  'align': 'center'}
    ]

    # ==================== 样式配置 ====================
    header_style = {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill("solid", fgColor="003366"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    }

    cell_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ==================== 数据处理 ====================
    def process_item(item):
        """强化数据处理，确保字段存在"""
        if item.get('eqtName') != '施工':
            return []

        main_info = {
            'cioName': item.get('cioName', ''),
            'eqtName': item.get('eqtName', ''),
            'csf': float(item.get('csf', 0)),
            'orgId': item.get('orgId', ''),
            'cecId': item.get('cecId', ''),
            'zzmx': ''  # 确保zzmx字段始终存在
        }

        details = item.get('zzmxcxfArray', [])
        if not details:
            # 返回带默认值的主信息
            return [main_info]

        processed = []
        for detail in details:
            processed.append({
                **main_info,
                'zzmx': detail.get('zzmx', ''),
                'cxdj': detail.get('cxdj', ''),
                'score': float(detail.get('score', 0)),
                'jcf': float(detail.get('jcf', 0)),
                'zxjf': float(detail.get('zxjf', 0)),
                'kf': float(detail.get('kf', 0)),
                'eqlId': detail.get('eqlId', '')
            })
        return processed

    # 生成基础数据
    processed_data = []
    for item in data:
        if isinstance(item, dict):
            processed_data.extend(process_item(item))

    # ==================== 创建主工作簿 ====================
    wb = Workbook()
    utc8_offset = timezone(timedelta(hours=8))
    timestamp = datetime.now(utc8_offset).strftime("%Y%m%d_%H%M%S")
    # ==================== 工作表配置 ====================
    sheet_configs = [
        {
            "name": "企业信用数据汇总",
            "prefix": None,
            "freeze": 'B2',
            "merge": True
        },
        {
            "name": "建筑工程总承包信用分排序",
            "prefix": "建筑业企业资质_施工总承包_建筑工程_",
            "freeze": 'B2',
            "merge": False,
            "generate_json": True  # 新增JSON生成标记
        },
        {
            "name": "市政公用工程信用分排序",
            "prefix": "建筑业企业资质_施工总承包_市政公用工程_",
            "freeze": 'B2',
            "merge": False,
            "generate_json": True  # 新增JSON生成标记
        },
        {
            "name": "装修装饰工程信用分排序",
            "prefix": "建筑业企业资质_专业承包_建筑装修装饰工程_",
            "freeze": 'B2',
            "merge": False,
            "generate_json": True  # 新增JSON生成标记
        },
        {
            "name": "水利水电工程信用分排序",
            "prefix": "建筑业企业资质_施工总承包_水利水电工程_",
            "freeze": 'B2',
            "merge": False,
            "generate_json": True  # 新增JSON生成标记
        },
        {
            "name": "电力工程信用分排序",
            "prefix": "建筑业企业资质_施工总承包_电力工程_",
            "freeze": 'B2',
            "merge": False,
            "generate_json": True  # 新增JSON生成标记
        }
    ]

    # ==================== 文件输出配置 ====================
    output_dir = os.getcwd()
    if github_mode:
        output_dir = os.path.join(output_dir, "excel_output")
        os.makedirs(output_dir, exist_ok=True)

    json_files = []
    detail_cache = {}  # 企业明细缓存

    # ==================== 构建各工作表 ====================
    # 先创建汇总表
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = sheet_configs[0]["name"]

    # 然后创建其他工作表
    for config in sheet_configs[1:]:
        ws = wb.create_sheet(title=config["name"])
        print(f"已创建工作表: {ws.title}")  # 调试日志

    # ==================== 填充各工作表 ====================
    for config in sheet_configs:
        # 获取工作表对象
        if config["name"] == "企业信用数据汇总":
            ws = summary_sheet
        else:
            ws = wb[config["name"]]

        print(f"\n正在处理工作表: {ws.title}")  # 调试日志

        # 设置冻结窗格
        ws.freeze_panes = config["freeze"]

        # ========== 写入表头 ==========
        headers = [col['name'] for col in COLUMNS]
        ws.append(headers)
        print(f"表头写入完成，行数: {ws.max_row}")  # 调试日志

        # 应用表头样式
        for col_idx, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx)
            for attr, value in header_style.items():
                setattr(cell, attr, value)
            ws.column_dimensions[get_column_letter(col_idx)].width = col['width']

        # ========== 处理数据 ==========
        if config["name"] == "企业信用数据汇总":
            sheet_data = processed_data
            merge_map = {}
        else:
            # 过滤排序数据
            sheet_data = sorted(
                [d for d in processed_data 
                 if str(d.get('zzmx', '')).startswith(config["prefix"]) 
                 and '级' in str(d.get('zzmx', ''))],
                key=lambda x: x.get('score', 0), 
                reverse=True
            )
            print(f"过滤到数据量: {len(sheet_data)}")  # 调试日志
            
            # ===== 新增：为所有前10名企业获取信誉分明细 =====
            # 遍历前10名企业
            for item in sheet_data[:10]:
                cec_id = item.get('cecId')
                company_name = item.get('cioName')
                
                if not cec_id:
                    print(f"警告: 企业 {company_name} 缺少cecId，跳过")
                    continue
                    
                # 检查缓存
                if cec_id in detail_cache:
                    item['detail'] = detail_cache[cec_id]
                    print(f"使用缓存获取企业信誉分明细: {company_name}")
                else:
                    # 获取企业信誉分明细
                    time.sleep(random.uniform(2, 5))
                    detail = fetch_company_detail(session, cec_id, company_name, max_retries=3) # 增强容错率
                    if detail:
                        item['detail'] = detail
                        detail_cache[cec_id] = detail
                    else:
                        print(f"警告: 未获取到企业 {company_name} 的信誉分明细")

            # 生成JSON（仅限指定工作表）
            if config.get("generate_json"):
                print(f"\n正在生成 {config['name']} 的JSON排行榜...")
                json_path = append_top_json(sheet_data, config["name"], github_mode)
                if json_path:
                    json_files.append(json_path)

        # ========== 写入数据 ==========
        if len(sheet_data) == 0:
            print(f"警告: {config['name']} 无数据，跳过写入")
            continue
        # ==========合并单元格逻辑（仅汇总表）==========
        current_key = None
        start_row = 2

        for row_idx, row_data in enumerate(sheet_data, 2):
            # 调试：打印前3行数据
            if row_idx <= 4:
                print(f"写入行 {row_idx} 数据: {row_data['zzmx'][:20]}...")

            # 企业信用数据汇总需要合并单元格
            if config["merge"]:
                unique_key = f"{row_data['orgId']}-{row_data['cecId']}"
                if unique_key != current_key:
                    if current_key is not None:
                        merge_map[current_key] = (start_row, row_idx-1)
                    current_key = unique_key
                    start_row = row_idx

            # 写入行数据
            row = [row_data.get(col['id'], '') for col in COLUMNS]
            ws.append(row)

            # ========== 设置数据单元格样式 ==========
            for col_idx in range(1, len(COLUMNS)+1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                col_def = COLUMNS[col_idx-1]
                
                # 修改点：始终设置垂直居中，保留原有水平对齐设置
                cell.alignment = Alignment(
                    horizontal=col_def['align'],  # 保留列定义的水平对齐方式
                    vertical='center',           # 新增垂直居中设置
                    wrap_text=False               # 可选：自动换行 True
                )
                
                if col_def.get('format'):
                    cell.number_format = col_def['format']

        # ========== 合并单元格（仅汇总表）==========
        if config["merge"]:
            # 处理最后一组
            if current_key:
                end_row = len(sheet_data) + 1
                if start_row <= end_row:
                    merge_map[current_key] = (start_row, end_row)

            # 执行合并
            for col in COLUMNS:
                if col['merge']:
                    col_letter = get_column_letter(COLUMNS.index(col)+1)
                    for (start, end) in merge_map.values():
                        if end > start:
                            ws.merge_cells(f"{col_letter}{start}:{col_letter}{end}")


    # ==================== 最终验证 ====================
    print("\n最终工作表列表:")
    for sheet in wb.sheetnames:
        print(f"- {sheet}")

    print(f"\n各工作表数据量:")
    for sheet in wb.worksheets:
        print(f"{sheet.title}: {sheet.max_row-1} 行")  # 减去表头

    # ==================== 文件保存 ====================
    filename = f"宜昌市信用评价信息_{timestamp}.xlsx" if github_mode else "宜昌市信用评价信息.xlsx"

    if github_mode:
        output_dir = os.path.join(os.getcwd(), "excel_output")
        # 确保目录创建成功
        try:
            os.makedirs(output_dir, exist_ok=True)
            print(f"已创建输出目录: {output_dir}")
        except Exception as e:
            print(f"目录创建失败: {str(e)}")
            raise

        filename = os.path.join(output_dir, filename)
        print(f"最终保存路径: {filename}")  # 路径调试

    try:
        # 删除默认创建的空白工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(filename)
        print(f"文件已保存至：{os.path.abspath(filename)}")
        print("包含的工作表:")
        for sheet in wb.sheetnames:
            print(f"- {sheet}")

        return {
            "excel": filename,
            "json": json_files
        }
    except Exception as e:
        print(f"文件保存失败：{str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    print("=== 启动数据获取程序 ===")
    session = requests.Session()
    all_data = []
    json_files = []

    try:
        # 初始获取验证码
        current_code, current_ts = get_new_code(session)
        print(f"[初始化] 验证码: {current_code} | 时间戳: {current_ts}")

        # 获取第一页确定总数
        first_data, total = process_page(session, 1, current_code, current_ts)
        total_pages = 100 #测试
        # total_pages = (total + PAGE_SIZE - 1) // PAGE_SIZE
        print(f"[初始化] 总记录数: {total} | 总页数: {total_pages}")

        if total == 0:
            print("错误: API返回总记录数为0，无需继续处理")
            return

        # 分页处理
        page = 1
        while page <= total_pages:
            retry_count = 0
            success = False

            while retry_count < PAGE_RETRY_MAX and not success:
                try:
                    print(f"\n[处理中] 第 {page} 页 (重试次数: {retry_count})")
                    page_data, _ = process_page(session, page, current_code, current_ts)

                    if page_data:
                        print(f"[成功获取数据] 第 {page} 页 {len(page_data)} 条记录")
                        all_data.extend(page_data)
                        success = True
                        page += 1
                    else:
                        print(f"[警告] 第 {page} 页获取到空数据，尝试刷新验证码")
                        raise RuntimeError("empty page data")

                except Exception as e:
                    retry_count += 1
                    print(f"[重试] 第 {page} 页第 {retry_count} 次重试: {str(e)}")

                    # 获取新验证码
                    try:
                        current_code, current_ts = get_new_code(session)
                        print(f"[刷新] 新验证码: {current_code} | 新时间戳: {current_ts}")
                    except Exception as e:
                        print(f"[警告] 验证码刷新失败: {str(e)}")
                        break

            if not success:
                print(f"[终止] 第 {page} 页超过最大重试次数，跳过此页")
                page += 1  # 跳过失败页

        print(f"\n=== 数据获取完成 ===")
        print(f"总获取记录数: {len(all_data)}")

        # 导出数据前再次检查
        if all_data:
            export_result = export_to_excel(all_data, session, github_mode=True)
            if export_result:
                json_files = export_result.get("json", [])
                
                # 设置GitHub Actions输出
                github_output = os.getenv('GITHUB_OUTPUT')
                if github_output:
                    # Excel文件路径
                    with open(github_output, 'a') as f:
                        f.write(f'excel-path={export_result["excel"]}\n')
                    
                    # JSON文件路径（多个）
                    for i, json_path in enumerate(json_files, 1):
                        with open(github_output, 'a') as f:
                            f.write(f'json-path-{i}={json_path}\n')
                else:
                    print("::注意:: 未在GitHub Actions环境中，跳过输出设置")
                
                # 打印所有生成的文件路径
                print("\n=== 所有生成的文件 ===")
                print(f"Excel文件: {export_result['excel']}")
                for i, json_path in enumerate(json_files, 1):
                    print(f"JSON文件 #{i}: {json_path}")
        else:
            print("错误: 没有获取到任何有效数据，无法导出Excel")
    except Exception as e:
        print(f"\n!!! 程序执行失败 !!!\n错误原因: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        session.close()

if __name__ == "__main__":
    main()
